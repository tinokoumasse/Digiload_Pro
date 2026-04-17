"""
Digiload Pro — Report Generator v1.0
Phase 4: PDF and Excel reports for mission proof of delivery

PDF:  WeasyPrint (HTML → PDF)
XLSX: openpyxl

Called from central_app.py:
    from reports import generate_pdf, generate_excel
"""

import os
import io
import json
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
# PDF REPORT
# ─────────────────────────────────────────────────────────────────────────────
def generate_pdf(mission: dict, pallets: list, clips: dict) -> bytes:
    """
    Generate a PDF proof of delivery for a mission.

    Args:
        mission: mission row dict from PostgreSQL
        pallets: list of pallet row dicts
        clips:   dict { sscc → filename } for clip links

    Returns:
        PDF bytes
    """
    from weasyprint import HTML

    loaded   = [p for p in pallets if p["status"] == "LOADED"]
    flagged  = [p for p in pallets if p["status"] == "FLAGGED"]
    waiting  = [p for p in pallets if p["status"] == "WAITING"]
    total    = len(pallets)
    pct      = round(len(loaded) / total * 100, 1) if total > 0 else 0

    activated  = _fmt_dt(mission.get("activated_at"))
    completed  = _fmt_dt(mission.get("completed_at"))
    generated  = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # Build pallet rows HTML
    pallet_rows = ""
    for i, p in enumerate(pallets, 1):
        status_color = {
            "LOADED":  "#30d158",
            "FLAGGED": "#ffd60a",
            "WAITING": "#7a90b8",
        }.get(p["status"], "#7a90b8")

        clip_link = ""
        if p["sscc"] in clips:
            clip_link = f'<a href="{clips[p["sscc"]]}" style="color:#2f7df6">▶ View</a>'

        pallet_rows += f"""
        <tr>
            <td>{i}</td>
            <td style="font-family:monospace">{p["sscc"]}</td>
            <td>{p.get("sku") or "—"}</td>
            <td><span style="color:{status_color};font-weight:600">{p["status"]}</span></td>
            <td>{_fmt_time(p.get("loaded_at"))}</td>
            <td>{f'#{p["forklift_id"]}' if p.get("forklift_id") is not None else "—"}</td>
            <td>{clip_link}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8"/>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: 'Helvetica Neue', Arial, sans-serif;
    font-size: 11px;
    color: #1a1a2e;
    background: #fff;
    padding: 32px;
  }}
  .header {{
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
    border-bottom: 3px solid #2f7df6;
    padding-bottom: 16px;
    margin-bottom: 24px;
  }}
  .logo {{
    font-size: 22px;
    font-weight: 900;
    color: #2f7df6;
    letter-spacing: 2px;
  }}
  .logo small {{
    display: block;
    font-size: 9px;
    font-weight: 400;
    letter-spacing: 4px;
    color: #7a90b8;
    margin-top: 2px;
  }}
  .doc-info {{
    text-align: right;
    color: #7a90b8;
    font-size: 10px;
    line-height: 1.8;
  }}
  .doc-info strong {{ color: #1a1a2e; }}
  h1 {{
    font-size: 18px;
    font-weight: 800;
    margin-bottom: 6px;
    color: #1a1a2e;
  }}
  .mission-meta {{
    display: grid;
    grid-template-columns: 1fr 1fr 1fr;
    gap: 12px;
    margin: 20px 0;
    padding: 16px;
    background: #f8f9ff;
    border-radius: 8px;
    border: 1px solid #e8ecff;
  }}
  .meta-item {{ display: flex; flex-direction: column; gap: 3px; }}
  .meta-label {{ font-size: 9px; text-transform: uppercase; letter-spacing: 1px; color: #7a90b8; }}
  .meta-value {{ font-size: 13px; font-weight: 700; color: #1a1a2e; }}
  .summary {{
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 10px;
    margin: 20px 0;
  }}
  .stat {{
    text-align: center;
    padding: 14px;
    border-radius: 8px;
    border: 1px solid #e8ecff;
  }}
  .stat-n {{
    font-size: 28px;
    font-weight: 900;
    line-height: 1;
    margin-bottom: 4px;
  }}
  .stat-l {{ font-size: 9px; text-transform: uppercase; letter-spacing: 1px; color: #7a90b8; }}
  .stat-n.blue {{ color: #2f7df6; }}
  .stat-n.green {{ color: #30d158; }}
  .stat-n.amber {{ color: #ffd60a; }}
  .stat-n.grey  {{ color: #7a90b8; }}
  .progress-bar {{
    height: 8px;
    background: #e8ecff;
    border-radius: 4px;
    overflow: hidden;
    margin: 16px 0;
  }}
  .progress-fill {{
    height: 100%;
    background: linear-gradient(90deg, #2f7df6, #30d158);
    border-radius: 4px;
    width: {pct}%;
  }}
  .progress-label {{
    font-size: 11px;
    color: #7a90b8;
    text-align: right;
    margin-bottom: 20px;
  }}
  table {{
    width: 100%;
    border-collapse: collapse;
    margin-top: 8px;
  }}
  thead th {{
    background: #2f7df6;
    color: #fff;
    font-size: 9px;
    font-weight: 700;
    letter-spacing: 1px;
    text-transform: uppercase;
    padding: 8px 10px;
    text-align: left;
  }}
  tbody tr:nth-child(even) {{ background: #f8f9ff; }}
  tbody tr:hover {{ background: #eef1ff; }}
  tbody td {{
    padding: 7px 10px;
    border-bottom: 1px solid #f0f0f0;
    font-size: 10px;
    color: #1a1a2e;
  }}
  .flagged-section {{
    margin-top: 24px;
    padding: 14px;
    background: #fff8e1;
    border: 1px solid #ffd60a;
    border-radius: 8px;
  }}
  .flagged-title {{
    font-size: 11px;
    font-weight: 700;
    color: #b8860b;
    margin-bottom: 8px;
    text-transform: uppercase;
    letter-spacing: 1px;
  }}
  .footer {{
    margin-top: 32px;
    padding-top: 16px;
    border-top: 1px solid #e8ecff;
    display: flex;
    justify-content: space-between;
    font-size: 9px;
    color: #7a90b8;
  }}
  .signature-box {{
    margin-top: 24px;
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 20px;
  }}
  .sig {{
    border-top: 1px solid #1a1a2e;
    padding-top: 6px;
    font-size: 9px;
    color: #7a90b8;
  }}
  @page {{
    margin: 20mm;
    @bottom-right {{
      content: "Page " counter(page) " of " counter(pages);
      font-size: 9px;
      color: #7a90b8;
    }}
  }}
</style>
</head>
<body>

<div class="header">
  <div>
    <div class="logo">DIGILOAD PRO<small>Proof of Delivery</small></div>
  </div>
  <div class="doc-info">
    <div>Generated: <strong>{generated}</strong></div>
    <div>Report type: <strong>Mission Completion</strong></div>
    <div>Gate: <strong>Gate {mission.get("gate_id","—")}</strong></div>
  </div>
</div>

<h1>{mission.get("name","Mission")}</h1>

<div class="mission-meta">
  <div class="meta-item">
    <span class="meta-label">Truck / Reference</span>
    <span class="meta-value">{mission.get("truck_id") or "—"}</span>
  </div>
  <div class="meta-item">
    <span class="meta-label">Activated</span>
    <span class="meta-value">{activated}</span>
  </div>
  <div class="meta-item">
    <span class="meta-label">Completed</span>
    <span class="meta-value">{completed}</span>
  </div>
  <div class="meta-item">
    <span class="meta-label">WMS Reference</span>
    <span class="meta-value">{mission.get("wms_mission_id") or "—"}</span>
  </div>
  <div class="meta-item">
    <span class="meta-label">Status</span>
    <span class="meta-value">{mission.get("status","—")}</span>
  </div>
  <div class="meta-item">
    <span class="meta-label">Source</span>
    <span class="meta-value">{(mission.get("source") or "csv").upper()}</span>
  </div>
</div>

<div class="summary">
  <div class="stat">
    <div class="stat-n blue">{total}</div>
    <div class="stat-l">Total Pallets</div>
  </div>
  <div class="stat">
    <div class="stat-n green">{len(loaded)}</div>
    <div class="stat-l">Loaded</div>
  </div>
  <div class="stat">
    <div class="stat-n amber">{len(flagged)}</div>
    <div class="stat-l">Flagged</div>
  </div>
  <div class="stat">
    <div class="stat-n grey">{len(waiting)}</div>
    <div class="stat-l">Not Loaded</div>
  </div>
</div>

<div class="progress-bar"><div class="progress-fill"></div></div>
<div class="progress-label">{pct}% completion rate</div>

<table>
  <thead>
    <tr>
      <th>#</th>
      <th>SSCC Barcode</th>
      <th>SKU</th>
      <th>Status</th>
      <th>Loaded At</th>
      <th>Forklift</th>
      <th>Clip</th>
    </tr>
  </thead>
  <tbody>
    {pallet_rows}
  </tbody>
</table>

{"" if not flagged else f'''
<div class="flagged-section">
  <div class="flagged-title">⚠ Flagged Events ({len(flagged)} pallets)</div>
  <p style="font-size:10px;color:#7a90b8">
    The following barcodes were scanned but could not be validated.
    Please review with your logistics team.
  </p>
  <ul style="margin-top:8px;padding-left:16px;font-size:10px;color:#b8860b">
    {"".join(f"<li style='margin-top:4px'>{p['sscc']}</li>" for p in flagged)}
  </ul>
</div>
'''}

<div class="signature-box">
  <div class="sig">Driver / Carrier Signature</div>
  <div class="sig">Warehouse Manager Signature</div>
</div>

<div class="footer">
  <div>Digiload Pro — Camera-confirmed loading validation</div>
  <div>Document generated automatically — {generated}</div>
</div>

</body>
</html>"""

    return HTML(string=html).write_pdf()


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL REPORT
# ─────────────────────────────────────────────────────────────────────────────
def generate_excel(mission: dict, pallets: list) -> bytes:
    """
    Generate an Excel report for a mission.
    Returns XLSX bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Mission Report"

    # ── Colors ────────────────────────────────────────────────────────────
    BLUE   = "2F7DF6"
    GREEN  = "30D158"
    AMBER  = "FFD60A"
    RED    = "FF453A"
    DARK   = "05080F"
    LIGHT  = "F0F4FF"
    WHITE  = "FFFFFF"
    GREY   = "7A90B8"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(bold=False, color="000000", size=11):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    thin = Border(
        left=Side(style="thin", color="E0E8FF"),
        right=Side(style="thin", color="E0E8FF"),
        bottom=Side(style="thin", color="E0E8FF"),
    )

    # ── Header block ──────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = "DIGILOAD PRO — Proof of Delivery"
    ws["A1"].font      = Font(bold=True, color=WHITE, size=16, name="Calibri")
    ws["A1"].fill      = fill(BLUE)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:G2")
    ws["A2"] = mission.get("name","Mission")
    ws["A2"].font      = Font(bold=True, color=DARK, size=13, name="Calibri")
    ws["A2"].fill      = fill(LIGHT)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 26

    # ── Mission meta ──────────────────────────────────────────────────────
    meta = [
        ("Gate",       str(mission.get("gate_id","—"))),
        ("Truck",      mission.get("truck_id") or "—"),
        ("WMS Ref",    mission.get("wms_mission_id") or "—"),
        ("Status",     mission.get("status","—")),
        ("Activated",  _fmt_dt(mission.get("activated_at"))),
        ("Completed",  _fmt_dt(mission.get("completed_at"))),
        ("Generated",  datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    row = 4
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = font(bold=True, color=GREY, size=10)
        ws.cell(row=row, column=2, value=value).font  = font(bold=True, size=10)
        row += 1

    # ── Summary stats ──────────────────────────────────────────────────────
    loaded  = sum(1 for p in pallets if p["status"] == "LOADED")
    flagged = sum(1 for p in pallets if p["status"] == "FLAGGED")
    total   = len(pallets)
    pct     = round(loaded / total * 100, 1) if total > 0 else 0

    stats_row = row + 1
    stats = [
        ("Total",    total,   BLUE),
        ("Loaded",   loaded,  GREEN),
        ("Flagged",  flagged, AMBER),
        ("Rate",     f"{pct}%", BLUE),
    ]
    for i, (lbl, val, color) in enumerate(stats):
        col = i + 1
        ws.cell(row=stats_row,   column=col, value=lbl).font  = font(color=GREY, size=9)
        ws.cell(row=stats_row+1, column=col, value=val).font  = Font(bold=True, color=color, size=20, name="Calibri")
        ws.cell(row=stats_row,   column=col).alignment        = Alignment(horizontal="center")
        ws.cell(row=stats_row+1, column=col).alignment        = Alignment(horizontal="center")
        ws.row_dimensions[stats_row+1].height = 32

    # ── Pallet table ──────────────────────────────────────────────────────
    table_start = stats_row + 4
    headers = ["#", "SSCC Barcode", "SKU", "Status", "Loaded At", "Forklift ID", "Weight (kg)"]
    col_widths = [5, 26, 16, 12, 18, 12, 12]

    for i, (h, w) in enumerate(zip(headers, col_widths)):
        col = i + 1
        cell = ws.cell(row=table_start, column=col, value=h)
        cell.font      = Font(bold=True, color=WHITE, size=10, name="Calibri")
        cell.fill      = fill(DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[table_start].height = 20

    status_colors = {"LOADED": GREEN, "FLAGGED": AMBER, "WAITING": "CCCCCC"}

    for i, p in enumerate(pallets):
        data_row = table_start + 1 + i
        row_fill = fill(LIGHT) if i % 2 == 0 else fill(WHITE)
        values   = [
            i + 1,
            p.get("sscc",""),
            p.get("sku") or "",
            p.get("status",""),
            _fmt_dt(p.get("loaded_at")),
            p.get("forklift_id"),
            p.get("weight_kg"),
        ]
        for j, val in enumerate(values):
            col  = j + 1
            cell = ws.cell(row=data_row, column=col, value=val)
            cell.fill      = row_fill
            cell.border    = thin
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if j != 1 else "left")
            cell.font      = font(size=10)
            if j == 3:   # Status column
                sc = status_colors.get(str(val),"CCCCCC")
                cell.font = Font(bold=True, color=sc, size=10, name="Calibri")

    # ── Auto-filter on table ──────────────────────────────────────────────
    ws.auto_filter.ref = (
        f"A{table_start}:{get_column_letter(len(headers))}"
        f"{table_start + len(pallets)}"
    )

    # ── Second sheet: raw data ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Raw Data")
    raw_headers = ["sscc","sku","status","scan_time","loaded_at","forklift_id","weight_kg"]
    for j, h in enumerate(raw_headers, 1):
        ws2.cell(row=1, column=j, value=h).font = font(bold=True)
    for i, p in enumerate(pallets, 2):
        for j, h in enumerate(raw_headers, 1):
            ws2.cell(row=i, column=j, value=p.get(h))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _fmt_dt(val) -> str:
    if not val:
        return "—"
    try:
        if hasattr(val, "strftime"):
            return val.strftime("%d/%m/%Y %H:%M")
        return str(val)[:16].replace("T"," ")
    except Exception:
        return str(val)

def _fmt_time(val) -> str:
    if not val:
        return "—"
    try:
        if hasattr(val, "strftime"):
            return val.strftime("%H:%M:%S")
        return str(val)[11:19]
    except Exception:
        return str(val)
